"""Generate docs/dashboard/autoresearch_equity.xlsx with:
  - Sheet 'Summary': all experiments ranked by composite, sortable/filterable
  - Sheet 'Equity Matrix': per-fold compounded equity per experiment + chart
  - Sheet 'Champion Detail': current global champion's equity curve + chart

Ships the chart embedded so users can open in Excel and see the equity
trajectory (strategy vs flat $1000 investment) across the 7 test folds.

Run as part of _sync_dashboard_to_docs.py; the dashboard links to the
xlsx via a download button on the equity-curve panel.
"""
from __future__ import annotations
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "autoresearchindexstock" / "autoresearch_results"
OUT = ROOT / "docs" / "index_stock_dashboard" / "autoresearch_equity.xlsx"

# ---- Load data ----
entries = [json.loads(l) for l in
           (SRC / "experiment_log.jsonl").read_text(encoding="utf-8").splitlines()
           if l.strip()]
best = json.loads((SRC / "best_config.json").read_text(encoding="utf-8"))

wb = Workbook()

# --- Styles ---
hdr_font = Font(bold=True, color="FFFFFF", size=10)
hdr_fill = PatternFill("solid", fgColor="1F6FEB")
thin = Side(border_style="thin", color="30363D")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center")

def stylise_header(ws, row, n_cols):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = border

# ---- Sheet 1: Summary ----
summary = wb.active
summary.title = "Summary"
cols = ["Exp#", "Backbone", "Description", "Status", "Composite",
        "Test Sharpe", "Val Sharpe", "Train Sharpe",
        "Test Return %", "Test Equity $", "Test Pos Folds",
        "Val Pos Folds", "Win Rate %", "IC", "Hit %",
        "Precision", "Recall", "F1", "MCC",
        "Elapsed s"]
summary.append(cols)
stylise_header(summary, 1, len(cols))
# Sort by composite descending so champions appear on top
entries_sorted = sorted(entries, key=lambda e: (e.get("composite") or -1e9), reverse=True)
for e in entries_sorted:
    summary.append([
        e.get("experiment_num"), e.get("backbone"),
        (e.get("description") or "")[:90], e.get("status"),
        e.get("composite"), e.get("sharpe"), e.get("val_sharpe"),
        e.get("train_sharpe"),
        e.get("return_pct"), e.get("equity"),
        e.get("test_pos_folds"), e.get("val_pos_folds"),
        e.get("win_rate"), e.get("ic"), e.get("hit"),
        e.get("precision"), e.get("recall"), e.get("f1"), e.get("mcc"),
        e.get("elapsed_sec"),
    ])
# Column widths
widths = [7, 14, 50, 9, 11, 11, 11, 12, 13, 14, 14, 14, 11, 8, 8, 10, 8, 8, 7, 10]
for i, w in enumerate(widths, 1):
    summary.column_dimensions[get_column_letter(i)].width = w
summary.freeze_panes = "A2"
summary.auto_filter.ref = summary.dimensions

# Conditional formatting: colour composite column
from openpyxl.formatting.rule import ColorScaleRule
summary.conditional_formatting.add(
    f"E2:E{summary.max_row}",
    ColorScaleRule(
        start_type="min", start_color="F85149",
        mid_type="num", mid_value=0, mid_color="FFFFFF",
        end_type="max", end_color="3FB950",
    ),
)

# ---- Sheet 2: Equity Matrix (chart data + line chart) ----
equity = wb.create_sheet("Equity Matrix")
# Only entries with per_window data
with_folds = [e for e in entries_sorted
              if e.get("per_window") and len(e["per_window"]) >= 1]
# Header
fold_hdr = ["Exp#", "Backbone", "Composite", "start",
            "fold_1", "fold_2", "fold_3", "fold_4", "fold_5", "fold_6", "fold_7"]
equity.append(fold_hdr)
stylise_header(equity, 1, len(fold_hdr))

START = 1000.0
for e in with_folds:
    windows = [w for w in (e.get("per_window") or []) if not w.get("skipped")]
    series = [START]
    cur = START
    for w in windows:
        cur *= (1 + (w.get("return_pct") or 0) / 100)
        series.append(cur)
    # Pad to 8 cols (start + 7 folds) if fewer than 7 folds
    while len(series) < 8:
        series.append(None)
    equity.append([e.get("experiment_num"), e.get("backbone"),
                    e.get("composite"), *series])

# Column widths
for i, w in enumerate([7, 14, 11, 10, 10, 10, 10, 10, 10, 10, 10], 1):
    equity.column_dimensions[get_column_letter(i)].width = w
equity.freeze_panes = "A2"

# Line chart — top 8 composite experiments' equity trajectories
chart = LineChart()
chart.title = "Equity curves — top 8 experiments (strategy vs $1000 flat investment)"
chart.style = 2
chart.y_axis.title = "Equity ($)"
chart.x_axis.title = "Fold index"
chart.height = 12
chart.width = 22
top_n = min(8, equity.max_row - 1)
# x axis categories (fold labels from header)
cats = Reference(equity, min_col=4, max_col=11, min_row=1, max_row=1)
for r in range(2, 2 + top_n):
    data = Reference(equity, min_col=4, max_col=11, min_row=r, max_row=r)
    chart.add_data(data, titles_from_data=False)
    series_title = f"Exp{equity.cell(row=r, column=1).value} ({equity.cell(row=r, column=2).value})"
    if chart.series:
        chart.series[-1].tx = None  # avoid autobuilt title
        from openpyxl.chart.series import SeriesLabel
        from openpyxl.chart.data_source import StrRef
        # Custom approach: set the series title via the first-column reference
chart.set_categories(cats)
# Place chart below the data
chart_anchor_row = equity.max_row + 3
equity.add_chart(chart, f"A{chart_anchor_row}")

# ---- Sheet 3: Champion Detail ----
champ = wb.create_sheet("Champion Detail")
champ.append(["AutoResearch GLOBAL CHAMPION"])
champ["A1"].font = Font(bold=True, size=14, color="58A6FF")
champ.append([])
champ.append(["Backbone", best.get("backbone")])
champ.append(["Experiment #", best.get("experiment_num")])
champ.append(["Description", best.get("description")])
champ.append(["Composite", best.get("composite")])
champ.append(["Test Sharpe", best.get("sharpe")])
champ.append(["Val Sharpe", best.get("val_sharpe")])
champ.append(["Test Return %", best.get("return_pct")])
champ.append(["Final Test Equity", best.get("equity")])
champ.append([])
champ.append(["Fold", "Regime", "Sharpe", "Return %", "Equity in fold",
               "Compounded Equity", "Investment Baseline"])
stylise_header(champ, 12, 7)
cur = START
for w in best.get("per_window", []):
    cur *= (1 + (w.get("return_pct") or 0) / 100)
    champ.append([
        w.get("fold"), w.get("regime"), w.get("sharpe"),
        w.get("return_pct"), w.get("equity"),
        round(cur, 2), START,
    ])
# Column widths
for i, w in enumerate([12, 32, 10, 11, 16, 18, 19], 1):
    champ.column_dimensions[get_column_letter(i)].width = w

# Chart — compounded equity vs flat investment
chart2 = LineChart()
chart2.title = f"Champion equity curve — {best.get('backbone')} Exp{best.get('experiment_num')}"
chart2.style = 12
chart2.y_axis.title = "Equity ($)"
chart2.x_axis.title = "Fold"
chart2.height = 12
chart2.width = 22
data_row_start = 13
data_row_end = 12 + len(best.get("per_window", []))
strat = Reference(champ, min_col=6, max_col=6, min_row=12, max_row=data_row_end)
invest = Reference(champ, min_col=7, max_col=7, min_row=12, max_row=data_row_end)
chart2.add_data(strat, titles_from_data=True)
chart2.add_data(invest, titles_from_data=True)
cats2 = Reference(champ, min_col=1, min_row=data_row_start, max_row=data_row_end)
chart2.set_categories(cats2)
champ.add_chart(chart2, f"I1")

# ---- Save ----
OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
size_kb = OUT.stat().st_size / 1024
print(f"Wrote {OUT} ({size_kb:.1f} KB, {len(entries)} experiments)")
